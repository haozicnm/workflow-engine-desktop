"""MCP Excel Server — openpyxl-based Excel operations.

Tools:
  excel_read    — Read data from a sheet
  excel_write   — Write data to a sheet (overwrite)
  excel_create  — Create a new workbook with headers
  excel_filter  — Filter in-memory data by column condition
  excel_sort    — Sort in-memory data by column
  excel_append  — Append rows to an existing file
  excel_csv     — Convert between CSV and Excel
"""

import json
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from mcp_protocol import McpServer, McpTool, log_stderr


class ExcelReadTool(McpTool):
    def __init__(self):
        super().__init__(
            name="excel_read",
            description="Read data from an Excel sheet",
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Path to .xlsx/.xls file"},
                    "sheet": {"type": "string", "description": "Sheet name (default: first sheet)"},
                },
                "required": ["path"],
            },
        )

    def execute(self, arguments: dict) -> dict:
        import openpyxl
        path = arguments["path"]
        sheet_name = arguments.get("sheet")
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()
        headers = rows[0] if rows else []
        data = []
        for r in rows[1:]:
            data.append(dict(zip(headers, r)) if headers else r)
        return {"headers": headers, "data": data, "rows": rows, "sheet": ws.title}


class ExcelWriteTool(McpTool):
    def __init__(self):
        super().__init__(
            name="excel_write",
            description="Write data to an Excel sheet (overwrite)",
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "sheet": {"type": "string", "default": "Sheet1"},
                    "data": {"type": "array", "description": "Array of arrays or objects"},
                    "write_mode": {"type": "string", "default": "overwrite"},
                },
                "required": ["path", "data"],
            },
        )

    def execute(self, arguments: dict) -> dict:
        import openpyxl
        from openpyxl.utils import get_column_letter
        path = arguments["path"]
        sheet_name = arguments.get("sheet", "Sheet1")
        data = arguments["data"]

        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        ws.title = sheet_name
        ws.delete_rows(1, ws.max_row)

        for i, row in enumerate(data):
            if isinstance(row, dict):
                if i == 0:
                    for j, k in enumerate(row.keys()):
                        ws.cell(row=1, column=j + 1, value=k)
                for j, v in enumerate(row.values()):
                    ws.cell(row=i + 2, column=j + 1, value=v)
            elif isinstance(row, (list, tuple)):
                for j, v in enumerate(row):
                    ws.cell(row=i + 1, column=j + 1, value=v)

        wb.save(path)
        wb.close()
        return {"path": path, "sheet": sheet_name, "rows_written": len(data)}


class ExcelCreateTool(McpTool):
    def __init__(self):
        super().__init__(
            name="excel_create",
            description="Create a new Excel workbook",
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "sheet": {"type": "string", "default": "Sheet1"},
                    "headers": {"type": "array", "description": "Column headers"},
                },
                "required": ["path"],
            },
        )

    def execute(self, arguments: dict) -> dict:
        import openpyxl
        path = arguments["path"]
        sheet_name = arguments.get("sheet", "Sheet1")
        headers = arguments.get("headers", [])
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        for j, h in enumerate(headers):
            ws.cell(row=1, column=j + 1, value=h)
        wb.save(path)
        wb.close()
        return {"path": path, "sheet": sheet_name, "headers": headers}


class ExcelFilterTool(McpTool):
    def __init__(self):
        super().__init__(
            name="excel_filter",
            description="Filter in-memory data by column condition",
            input_schema={
                "type": "object",
                "properties": {
                    "data": {"type": "array"},
                    "column": {"type": "string"},
                    "op": {"type": "string", "description": "==, !=, <, >, <=, >=, contains, startswith"},
                    "value": {},
                },
                "required": ["data", "column", "op", "value"],
            },
        )

    def execute(self, arguments: dict) -> dict:
        data = arguments["data"]
        column = arguments["column"]
        op = arguments["op"]
        value = arguments["value"]

        ops = {
            "==": lambda a, b: a == b,
            "!=": lambda a, b: a != b,
            "<": lambda a, b: float(str(a)) < float(str(b)),
            ">": lambda a, b: float(str(a)) > float(str(b)),
            "<=": lambda a, b: float(str(a)) <= float(str(b)),
            ">=": lambda a, b: float(str(a)) >= float(str(b)),
            "contains": lambda a, b: str(b).lower() in str(a).lower(),
            "startswith": lambda a, b: str(a).lower().startswith(str(b).lower()),
        }
        op_fn = ops.get(op, ops["=="])
        filtered = [row for row in data if isinstance(row, dict) and column in row and op_fn(row[column], value)]
        return {"data": filtered, "count": len(filtered)}


class ExcelSortTool(McpTool):
    def __init__(self):
        super().__init__(
            name="excel_sort",
            description="Sort in-memory data by column",
            input_schema={
                "type": "object",
                "properties": {
                    "data": {"type": "array"},
                    "column": {"type": "string"},
                    "order": {"type": "string", "default": "asc"},
                },
                "required": ["data", "column"],
            },
        )

    def execute(self, arguments: dict) -> dict:
        data = arguments["data"]
        column = arguments["column"]
        order = arguments.get("order", "asc")
        reverse = order == "desc"
        sorted_data = sorted(data, key=lambda r: r.get(column, "") if isinstance(r, dict) else r, reverse=reverse)
        return {"data": sorted_data}


class ExcelAppendTool(McpTool):
    def __init__(self):
        super().__init__(
            name="excel_append",
            description="Append rows to an existing Excel file",
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "sheet": {"type": "string", "default": "Sheet1"},
                    "data": {"type": "array"},
                },
                "required": ["path", "data"],
            },
        )

    def execute(self, arguments: dict) -> dict:
        import openpyxl
        path = arguments["path"]
        sheet_name = arguments.get("sheet", "Sheet1")
        data = arguments["data"]
        if not os.path.exists(path):
            return ExcelWriteTool().execute(arguments)
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        start_row = ws.max_row + 1
        for i, row in enumerate(data):
            if isinstance(row, dict):
                values = list(row.values())
            elif isinstance(row, (list, tuple)):
                values = list(row)
            else:
                values = [row]
            for j, v in enumerate(values):
                ws.cell(row=start_row + i, column=j + 1, value=v)
        wb.save(path)
        wb.close()
        return {"path": path, "sheet": sheet_name, "rows_appended": len(data)}


class ExcelCsvTool(McpTool):
    def __init__(self):
        super().__init__(
            name="excel_csv",
            description="Convert between CSV and Excel",
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Input file path"},
                    "direction": {"type": "string", "description": "csv_to_xlsx or xlsx_to_csv"},
                    "output": {"type": "string", "description": "Output file path"},
                    "delimiter": {"type": "string", "default": ","},
                },
                "required": ["path", "direction", "output"],
            },
        )

    def execute(self, arguments: dict) -> dict:
        import openpyxl, csv
        path = arguments["path"]
        direction = arguments["direction"]
        output = arguments["output"]
        delimiter = arguments.get("delimiter", ",")

        if direction == "csv_to_xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active
            with open(path, "r", newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f, delimiter=delimiter)
                for i, row in enumerate(reader):
                    for j, val in enumerate(row):
                        ws.cell(row=i + 1, column=j + 1, value=val)
            wb.save(output)
            wb.close()
        elif direction == "xlsx_to_csv":
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            with open(output, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f, delimiter=delimiter)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(list(row))
            wb.close()
        else:
            return {"error": f"Unknown direction: {direction}"}
        return {"path": output, "direction": direction}


def main():
    server = McpServer("wf-excel", "1.0.0")
    server.add_tool(ExcelReadTool())
    server.add_tool(ExcelWriteTool())
    server.add_tool(ExcelCreateTool())
    server.add_tool(ExcelFilterTool())
    server.add_tool(ExcelSortTool())
    server.add_tool(ExcelAppendTool())
    server.add_tool(ExcelCsvTool())
    server.run()


if __name__ == "__main__":
    main()
